import os
import pyzipper
import rarfile # type: ignore
import py7zr # type: ignore
from datetime import datetime
import configparser
import hashlib
import uuid
import re
import yaml

from openpyxl import load_workbook
import csv

from datetime import datetime
from elasticsearch import Elasticsearch # type: ignore

config = configparser.ConfigParser()
config.read('config.ini')

ELASTIC_URL = config['ELASTIC_API']['ELASTIC_URL']
ELASTIC_API_KEY = config['ELASTIC_API']['ELASTIC_API_KEY']
ELASTIC_API_CACERT = config['ELASTIC_API']['ELASTIC_API_CACERT']

history_read = './history_read.txt'
history_downloaded = './history_downloaded.txt'



async def get_file_hash_before_download(client, message):
    document = message.media.document
            
    # Tạo buffer để lưu toàn bộ dữ liệu
    file_data = bytearray()
    
    # Tải và lưu từng chunk
    async for chunk in client.iter_download(document, chunk_size=1024*1024):
        file_data.extend(chunk)
    
    # Tính hash từ data trong memory
    return hashlib.md5(file_data).hexdigest()

def get_file_hash_after_download(file_path, chunk_size=1024*1024):
    md5_hash = hashlib.md5()
    total_bytes = 0
    with open(file_path, "rb") as f:
        for byte_block in iter(lambda: f.read(chunk_size), b""):
            md5_hash.update(byte_block)
            total_bytes += len(byte_block)
    return md5_hash.hexdigest()

# Hàm thêm line vào file
def append_line_to_file(history_file, file_name):
    try:
        with open(history_file, 'a', encoding='utf-8') as f:
            f.write(f'{file_name}\n')

    except Exception as e:
        print(f'Error marking file download: {str(e)}')


# Hàm kiểm tra đoạn chat có phải là nhóm hoặc channel không
def check_chat_type(chat_id):
    if str(chat_id).startswith('-100'):
        return True
    return False


# Hàm kiểm tra định dạng file có phải file nén không
compress_extensions = ['.zip', '.rar', '.7z']
def check_compress_file(file):
    _, file_extension = os.path.splitext(file)
    return file_extension in compress_extensions


# Hàm kiểm tra mail có thuộc miền chỉ định không, nếu không chỉ định email nào thì mặc định là mọi miền 
mail_extensions = []
def check_custom_mail(mail):
    if len(mail_extensions) == 0:
        return True
    return any(mail.endswith(suffix) for suffix in mail_extensions)


# Hàm kiểm tra xem tên file đã tồn tại trong file history chưa
def check_file_in_history(history_file, file_hash):
    try:
        # Kiểm tra file có tồn tại không
        if not os.path.exists(history_file):
            return False

        # Mở và đọc file
        with open(history_file, 'r', encoding='utf-8') as f:
            content = f.read()
            if file_hash in content:
                return True
    except Exception as e:
        print(f'Error checking file existence: {str(e)}')
    return False


# Hàm kiểm tra xem file có đúng định dạng được chỉ định không, nếu không chỉ định extension nào sẽ chấp nhận mọi kiểu file
file_extensions = ['.txt', '.zip', '.7z', '.rar']
def check_valid_file_extension(file):
    _, file_extension = os.path.splitext(file)
    return file_extension in file_extensions

# Hàm Rename tên file sau khi giải nén
async def rename_extract_file(extract_file_path, new_files):
    renamed_files = []
    for item in new_files:
        item_path = os.path.join(extract_file_path, item)

        if os.path.isfile(item_path):  # Nếu là file
            file_size = str(os.path.getsize(item_path)) 
            # Đổi tên file
            new_name = f'{file_size}-{item_path}'
            new_path = os.path.join(extract_file_path, new_name)
            os.rename(item_path, new_path)
            renamed_files.append(new_name)
        
        elif os.path.isdir(item_path):  # Nếu là thư mục
            # Đổi tên các file bên trong thư mục
            for root, _, files in os.walk(item_path):
                for file_name in files:
                    old_file_path = os.path.join(root, file_name)
                    # Lấy kích thước file
                    file_size = os.path.getsize(old_file_path)
                    # Tạo tên file mới
                    new_file_name = f'{file_size}-{file_name}'

                    new_file_path = os.path.join(root, new_file_name)
                    os.rename(old_file_path, new_file_path)
                    renamed_files.append(os.path.relpath(new_file_path, extract_file_path))

    return renamed_files

async def extract_file(file_path, extract_dir, password=None):
    try:
        os.makedirs(extract_dir, exist_ok=True)
        _, file_extension = os.path.splitext(file_path)
        before_extract = set(os.listdir(extract_dir))

        if (file_extension == '.zip'):   
            with pyzipper.AESZipFile(file_path, 'r') as zip_ref:
                if password:
                    zip_ref.extractall(extract_dir, pwd=password.encode('utf-8'))
                else:
                    zip_ref.extractall(extract_dir)
        if (file_extension == '.rar'):
            with rarfile.RarFile(file_path, 'r') as rar_ref:
                if password:
                    rar_ref.extractall(extract_dir, pwd=password)
                else:
                    rar_ref.extractall(extract_dir)
        if (file_extension == '.7z'):
            with py7zr.SevenZipFile(file_path, mode='r', password=password) as archive:
                archive.extractall(path=extract_dir)

        after_extract = set(os.listdir(extract_dir))

        new_files = after_extract - before_extract

        # Đổi tên file 
        renamed_files = []
        for item in new_files:
            item_path = os.path.join(extract_dir, item)

            if os.path.isfile(item_path):  # Nếu là file
                file_hash = get_file_hash_after_download(item_path)
                # Đổi tên file
                new_name = f'{file_hash}-{item_path}'
                new_path = os.path.join(extract_dir, new_name)
                os.rename(item_path, new_path)
                renamed_files.append(new_name)
            
            elif os.path.isdir(item_path):  # Nếu là thư mục
                # Đổi tên các file bên trong thư mục
                root_dir = None
                for root, _, files in os.walk(item_path):
                    for file_name in files:
                        old_file_path = os.path.join(root, file_name)
                        # Lấy mã băm file
                        file_hash = get_file_hash_after_download(old_file_path)
                        # Tạo tên file mới
                        new_file_name = f'{file_hash}-{file_name}'

                        new_file_path = os.path.join('./storage', new_file_name)
                        os.rename(old_file_path, new_file_path)
                        renamed_files.append(os.path.relpath(new_file_path, extract_dir))
                    root_dir = root
                os.rmdir(root_dir)


        for file_path in renamed_files:
            if check_file_in_history(history_read, file_hash):
                print(f'Extract file zip file_path: {file_path}')
                os.remove(f'./storage/{file_path}')
        return list(renamed_files)
        
    except Exception as e:
        print(f'Error: {str(e)}')
    return None

# Hàm callback trả về % quá trình tải file
async def progress_callback(current, total):
    percentage = (current / total) * 100
    print(f'{current}/{total} bytes ({percentage:.2f}%)')


# Hàm tải file 
async def download_file_from_media(client, message, download_dir, file_hash):
    try:
        file_name = None
        if hasattr(message.media, 'document') and message.media.document:
            for attribute in message.media.document.attributes:
                if hasattr(attribute, 'file_name'):
                    file_name = str(file_hash) + '-' +  attribute.file_name
                    break
        print(f'File name: {file_name}')
        file_path = None
        if check_valid_file_extension(file_name):
            download_path = os.path.join(download_dir, file_name)
            # Bắt đầu tải file từ message 
            file_path = await client.download_media(
                message.media,
                file=download_path,
                progress_callback=progress_callback
            )
            if file_path is not None:
                append_line_to_file(history_downloaded, file_path)
        return file_path
            
    except Exception as e:
        print(f'Error during download: {str(e)}')
        import traceback
        print('Full error:', traceback.format_exc())
        return None
    
# Hàm kiểm tra định dạng của line
def check_line_format(rules, line):

    # Duyệt qua từng quy tắc trong danh sách
    for rule in rules:
        for rule_name, rule_pattern in rule.items():
            x = re.match(rule_pattern, line) 
            if x:
                return [rule_name, x]  # Trả về tên quy tắc đã khớp
    
    print('Not matching rule')
    return None  # Không có quy tắc nào khớp

def load_rules_from_yaml(rule_path):
    with open(rule_path, 'r') as file:
        rules = yaml.safe_load(file)
    return rules['line_rules']

link_rules='./rules/link_rules.yaml'
async def get_room_link_from_message(message):
    url = message['media']['webpage']['url']
    print(f"URL: {url}")
    rules = load_rules_from_yaml(link_rules)
    matches = check_line_format(rules, url)
    if matches:
        print(matches.group(1))
    

# Hàm đọc file
async def read_file(file_path):
    try:
        result = None
        elastic_client = Elasticsearch(
            [ELASTIC_URL],
            api_key=ELASTIC_API_KEY,  # Thay bằng API key bạn vừa tạo
            verify_certs=True,
            ca_certs=ELASTIC_API_CACERT
        )
        if os.path.isfile(file_path):

            if not check_compress_file(file_path):
                if not check_valid_file_extension(file_path):
                    return None
                file_hash = get_file_hash_after_download(file_path)
                if check_file_in_history(history_read, file_hash):
                    return None
                _, file_extension = os.path.splitext(file_path)
                if file_extension == '.txt':
                        with open(file_path, 'r') as file_txt:
                            for line in file_txt:
                                line = line.strip()
                                doc = {
                                    "text": line,
                                    "timestamp": datetime.now(),
                                }
                                try:
                                    response_elastic = elastic_client.index(index="telegram_index", id=str(uuid.uuid4()), document=doc)
                                    print(f'Response elastic: {response_elastic}')
                                except Exception as e:
                                    print(f"Error indexing document: {e}")
                 
                        append_line_to_file(history_read, file_path)
                        result = True
                        os.remove(file_path)
            else:  
                # Đường dẫn giải nén
                extract_dir = './storage/'
                list_extract = await extract_file(file_path, extract_dir)
                if list_extract:

                    # Sau khi giải nén, đọc nội dung các file trong thư mục đã giải nén
                    for extract_path_item in list_extract:
                        item_path = os.path.join(extract_dir, extract_path_item)
                        print(f'Item path: {item_path}')
                        if await read_file(item_path):
                            os.remove(item_path)
                        

                # Xóa file sau khi đã đọc và xử lý
                os.remove(file_path)
                print(f'File {file_path} has been processed and deleted.')

        elif os.path.isdir(file_path):
            for item in os.listdir(file_path):
                item_path = os.path.join(file_path, item)
                await read_file(item_path)
                result = True
            os.remove(file_path)
            print(f'File {file_path} has been processed and deleted.')
    
    except Exception as e:
        print(f'Error: {e}')
    finally:
        elastic_client.close()
        return result
# Nhóm hàm đọc file 
# Đọc file xlsx
def read_table_xlsx(path):
    workbook = load_workbook(filename=path, data_only=True)
    sheet = workbook.active  # Hoặc bạn có thể chỉ định sheet cụ thể

    columns = [cell.value for cell in sheet[1]]

    require_columns = ['CUSTOMER_NAME', 'BIRTH_DATE', 'PHONE', 'EMAIL', 'PASSWORD']
    valid_columns = [col for col in columns if col in require_columns]

    if valid_columns:
    
        for row in sheet.iter_rows(min_row=2, values_only=True): 
            row_data = {col: row[i] for i, col in enumerate(columns) if col in valid_columns}

            if row_data['EMAIL'] is not None:
                if check_custom_mail(row_data['EMAIL']):
                    if 'BIRTH_DATE' in row_data and isinstance(row_data['BIRTH_DATE'], datetime):
                        row_data['BIRTH_DATE'] = row_data['BIRTH_DATE'].strftime('%d/%m/%Y')
                    print(row_data)
    else:
        print('Không có cột nào trong danh sách yêu cầu tồn tại trong file.')

# Đọc file csv
def read_table_csv(path):
    with open(path, mode='r') as file:
        reader = csv.reader(file)
        for row in reader:
            print(row)


# Hàm lưu dữ liệu
async def save_data(channel, filename, url, user, password, name, phone):
    try:
        if user != '' and password != '':
            document = {'channel': channel, 'filename': filename, 'user': user, 'pass': password}
            if url != '':
                document['url'] = url 
            if name != '':
                document['name'] = name
            if phone != '':
                document['phone'] = phone
            

    except Exception as e:
        print(f'Error: {e}')

